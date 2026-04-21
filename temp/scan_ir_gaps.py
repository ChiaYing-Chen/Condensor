import openpyxl

wb = openpyxl.load_workbook('TG-1冷凝器銅管管板.xlsx', data_only=True)
sheet = wb.worksheets[0]

# Scan specifically the bottom half (IR/IL area, rows 52-105)
# Focus on the right half (cols 100-200) and check gap patterns
print("=== Bottom Half Right Side: Row 52-100, Col 99-200 ===")

TUBE_COLORS = {'FF00B0F0', 'FF0070C0'}  # 亮藍 + 深藍

for r in range(52, 101):
    row_summary = []
    in_tube = False
    gap_start = None
    for c in range(99, 201):
        cell = sheet.cell(row=r, column=c)
        color = '00000000'
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            color = str(cell.fill.fgColor.rgb)
        
        is_tube = color in TUBE_COLORS
        
        if is_tube and not in_tube:
            if gap_start is not None:
                row_summary.append(f'GAP[{gap_start}-{c-1}]')
            in_tube = True
        elif not is_tube and in_tube:
            row_summary.append(f'T[{c-1}]')  # last tube col
            gap_start = c
            in_tube = False
    
    # Count total tubes in this row (right half from col 99)
    cnt = 0
    for c in range(99, 201):
        cell = sheet.cell(row=r, column=c)
        color = '00000000'
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            color = str(cell.fill.fgColor.rgb)
        if color in TUBE_COLORS:
            cnt += 1
    
    # Check for internal gaps (large gaps > 3 cols)
    has_big_gap = any('GAP' in s and int(s.split('[')[1].split('-')[1].split(']')[0]) - int(s.split('[')[1].split('-')[0]) > 3 for s in row_summary if 'GAP' in s)
    
    marker = ' *** GAP ***' if has_big_gap else ''
    print(f"Row {r:3d}: {cnt:3d} colored cells{marker}")
    if has_big_gap:
        print(f"         {row_summary}")
