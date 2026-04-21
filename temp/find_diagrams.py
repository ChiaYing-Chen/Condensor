import openpyxl

wb = openpyxl.load_workbook('TG-1冷凝器銅管管板.xlsx', data_only=True)
sheet = wb.worksheets[0]

print("Scanning for text to find the second diagram...")
for r in range(1, 500):
    for c in range(1, 50):
        val = sheet.cell(row=r, column=c).value
        if val and isinstance(val, str) and ("原廠" in val or "氨腐蝕" in val or "TG-3" in val or "空冷" in val):
            print(f"Found text at Row {r}, Col {c}: {val}")

# Let's count colored cells row by row up to row 300
counts = []
for r in range(1, 300):
   cnt = 0
   for c in range(1, 300):
       cell = sheet.cell(row=r, column=c)
       if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
           color = str(cell.fill.fgColor.rgb)
           if color not in ['00000000', 'FFFFFFFF', '00FFFFFF', 'FF000000', 'FFD9D9D9', 'FFBFBFBF', 'FFA6A6A6']:
               cnt += 1
   counts.append(cnt)

print("Colored tubes per row (every 10 rows):")
for i in range(0, 300, 10):
   print(f"Row {i} to {i+9}: {sum(counts[i:i+10])}")
