import openpyxl
import json

def analyze_sheets():
    print("Loading workbook...")
    wb = openpyxl.load_workbook('TG-1冷凝器銅管管板.xlsx', data_only=True)
    
    TUBE_COLORS = ['FF00B0F0', 'FF0070C0']
    
    results = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Analyzing sheet: {sheet_name}")
        tube_count = 0
        rows_with_tubes = 0
        
        # Scan a larger area to be safe
        for r in range(1, 400):
            row_has_tube = False
            for c in range(1, 400):
                cell = sheet.cell(row=r, column=c)
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    color = str(cell.fill.fgColor.rgb)
                    if color in TUBE_COLORS:
                        tube_count += 1
                        row_has_tube = True
            if row_has_tube:
                rows_with_tubes += 1
        
        results[sheet_name] = {
            "tube_count": tube_count,
            "rows_with_tubes": rows_with_tubes
        }
        print(f"  -> Found {tube_count} tubes across {rows_with_tubes} rows.")

    with open('sheet_analysis.json', 'w') as f:
        json.dump(results, f, indent=2)

if __name__ == "__main__":
    analyze_sheets()
