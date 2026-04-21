import openpyxl
import json

def extract_shape():
    print("Loading workbook...")
    wb = openpyxl.load_workbook('TG-1冷凝器銅管管板.xlsx', data_only=True)
    
    print("Sheet names:", wb.sheetnames)
    
    # Just use the first sheet for visual layout
    sheet = wb.worksheets[0]
    print("Using sheet:", sheet.title)
    
    print("Extracting cell colors...")
    grid = []
    
    tubes_found = 0
    # Let's increase the search area just in case it grew
    for row in range(1, 150):
        row_data = []
        for col in range(1, 300):
            cell = sheet.cell(row=row, column=col)
            has_fill = False
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                color = str(cell.fill.fgColor.rgb)
                # Ignore common empty backgrounds
                if color not in ['00000000', 'FFFFFFFF', '00FFFFFF', 'FF000000', 'FFD9D9D9', 'FFBFBFBF', 'FFA6A6A6']:
                    has_fill = True
            
            row_data.append(1 if has_fill else 0)
            if has_fill: tubes_found += 1
            
        grid.append(row_data)
        
    print(f"Total tubes strictly filtered: {tubes_found}")
    
    with open('grid_shape_strict.json', 'w') as f:
        json.dump(grid, f)

if __name__ == "__main__":
    extract_shape()
