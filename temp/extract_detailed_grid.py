import openpyxl
import json

def extract_detailed_grid():
    print("Loading workbook...")
    wb = openpyxl.load_workbook('TG-1冷凝器銅管管板.xlsx', data_only=True)
    sheet = wb.worksheets[0]
    
    TUBE_COLORS = ['FF00B0F0', 'FF0070C0']
    
    # We found 99 rows contain tubes. Let's find the specific range.
    min_row, max_row = 999, 0
    min_col, max_col = 999, 0
    
    tubes = []
    
    for r in range(1, 400):
        for c in range(1, 400):
            cell = sheet.cell(row=r, column=c)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                color = str(cell.fill.fgColor.rgb)
                if color in TUBE_COLORS:
                    tubes.append({"r": r, "c": c, "color": color})
                    min_row = min(min_row, r)
                    max_row = max(max_row, r)
                    min_col = min(min_col, c)
                    max_col = max(max_col, c)

    print(f"Stats: Rows {min_row}-{max_row}, Cols {min_col}-{max_col}")
    print(f"Total tubes: {len(tubes)}")
    
    # Build a tight grid for analysis
    grid = []
    for r in range(min_row, max_row + 1):
        row_data = []
        for c in range(min_col, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            is_tube = 0
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                if str(cell.fill.fgColor.rgb) in TUBE_COLORS:
                    is_tube = 1
            row_data.append(is_tube)
        grid.append(row_data)

    output = {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
        "grid": grid
    }
    
    with open('detailed_grid.json', 'w') as f:
        json.dump(output, f)
    
    print("Saved detailed_grid.json")

if __name__ == "__main__":
    extract_detailed_grid()
