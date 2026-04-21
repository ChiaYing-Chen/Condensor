import openpyxl

wb = openpyxl.load_workbook('TG-1冷凝器銅管管板.xlsx', data_only=True)
print("Sheet names in workbook:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n--- Scanning Sheet: {sheet_name} ---")
    
    found_text = False
    for r in range(1, 200):
        for c in range(1, 50):
            val = sheet.cell(row=r, column=c).value
            if val and isinstance(val, str) and ("原廠" in val or "空冷" in val or "TG-3" in val):
                print(f"Found text at Row {r}, Col {c}: {val}")
                found_text = True
    
    if not found_text:
        print("No label text found in this sheet.")
